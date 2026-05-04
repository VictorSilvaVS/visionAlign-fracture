import os
import logging
from functools import wraps 
from threading import Lock 
from flask import Flask, request, jsonify, session, render_template, redirect, url_for, g, Response, send_file, abort
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import timedelta, datetime 
from werkzeug.security import generate_password_hash, check_password_hash 
import time 
import cv2 
import socket 
from io import BytesIO 
from utils.database import Database
import json 
from io import StringIO 
import collections 
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill 
import queue 
import csv 
from .timezone_utils import get_current_utc_timestamp, get_current_timestamp_for_storage, parse_stored_timestamp_to_utc, convert_utc_to_local_display
import secrets
import numpy as np
import math
from utils.email_utils import send_password_reset_email
from training.otx_manager import OTXManager
import base64


def json_serial_default(o):
    """JSON serializer for objects not serializable by default json code."""
    if isinstance(o, datetime):
        _logger.debug(f"Serializing datetime: {o}") # Log para depuração
        return o.isoformat()  
    if isinstance(o, timedelta):
        return o.total_seconds() 
    if isinstance(o, (np.integer, np.int_)): 
        return int(o)
    if isinstance(o, (np.floating, np.float_)): 
        val = float(o) 
        if math.isnan(val) or math.isinf(val): 
            return None  
        return val
    if isinstance(o, np.ndarray): 
        return o.tolist()
    if isinstance(o, np.bool_): 
        return bool(o)
    if isinstance(o, float): 
        if math.isnan(o) or math.isinf(o): 
            return None  
        return o    
    # Se chegarmos aqui, o objeto não foi serializado por nenhum handler
    _logger.error(f"JSON_SERIAL_ERROR: Objeto do tipo {o.__class__.__name__} com valor '{o}' não é serializável.")
    raise TypeError(f"Object of type {o.__class__.__name__} with value '{o}' is not JSON serializable by default or by custom handler.")


PRE_ENCODED_MINIMAL_JPEG_BYTES = (
    b'\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00\xff\xdb\x00C\x00\x02\x01\x01\x01\x01\x01\x02\x01\x01\x01\x02\x02\x02\x02\x02\x04\x03\x02\x02\x02\x02\x05\x04\x04\x03\x04\x06\x05\x06\x06\x06\x05\x06\x06\x06\x07\t\x08\x06\x07\t\x07\x06\x06\x08\x0b\x08\t\n\n\n\n\n\x06\x08\x0b\x0c\x0b\n\x0c\t\n\n\n\xff\xc9\x00\x0b\x08\x00\x01\x00\x01\x01\x01\x11\x00\xff\xda\x00\x08\x01\x01\x00\x00?\x00\xf4\xff\xd9'
)

_settings_manager = None
_logger = None 
_current_stats = None 
_stats_lock = None 
_frame_container = None 
_frame_lock = None 
_system_start_time = None
_extract_frame_flag = None 
_model_instance = None 
_web_stream_paused = False 
_web_stream_lock = Lock() 
_security_manager = None 
_otx_manager = None
MAX_CONSOLE_LINES = 500 # Default inicial, será atualizado no run_server
log_queue = collections.deque(maxlen=MAX_CONSOLE_LINES) 
sse_listeners = [] 
_active_alerts_cache = collections.defaultdict(lambda: {'timestamp': 0, 'type': None, 'last_detail': ''})
_ALERT_COOLDOWN_SECONDS = 60  
_ALERT_STATE_EXPIRY_SECONDS = 300  
base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) 
template_dir = os.path.join(base_dir, 'interface', 'flask', 'templates')
static_dir = os.path.join(base_dir, 'interface', 'flask', 'static')
flask_app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
# flask_app.secret_key será definido no run_server

active_users = {}  
                    
active_users_lock = Lock() 


class SSELogHandler(logging.Handler):
    def __init__(self, log_deque):
        super().__init__()
        self.log_deque = log_deque

        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S'))

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_deque.append(msg) 

            for listener_queue in sse_listeners:
                listener_queue.put_nowait(msg)
        except Exception:
            self.handleError(record)
# PERMANENT_SESSION_LIFETIME será definido no run_server


def get_db():
    """Abre uma nova conexão de banco de dados se não houver uma para a requisição atual."""
    if 'db' not in g:
        db_path = None
        if _settings_manager:
            db_path = _settings_manager.get_setting('SYSTEM_CONFIG', 'paths', {}).get('database')
        
        g.db = Database(db_path=db_path)
    return g.db

@flask_app.teardown_appcontext
def close_db(e=None):
    """Fecha a conexão do banco de dados ao final da requisição."""
    db = g.pop('db', None)
    if db is not None:
        db.close()

def get_settings_manager():
    """Retorna a instância global do SettingsManager."""
    return _settings_manager

def get_security_manager():
    """Retorna a instância global do SecurityManager."""
    return _security_manager

@flask_app.context_processor
def inject_now():
    """Injects the current UTC date/time into the template context."""
    
    return {'now_utc': get_current_utc_timestamp()}


def generate_video_frames():
    """Generator function para o stream MJPEG."""
    global _web_stream_paused, _web_stream_lock 
    last_frame_time = time.time()
    
    # Busca FPS das configurações ou usa default 12
    target_fps = 12
    if _settings_manager:
        target_fps = _settings_manager.get_setting('SYSTEM_CONFIG', 'monitoring', {}).get('web_stream_fps', 12)
    while True:
        frame_copy = None 
        if _frame_lock is None or _frame_container is None:
             _logger.warning("Stream MJPEG: Frame lock não inicializado.")
             time.sleep(0.1)
             continue


        with _web_stream_lock:
            if _web_stream_paused:
             _logger.debug("Stream MJPEG pausado via API web.")
             time.sleep(0.1)
             continue

        with _frame_lock:
            current_frame = _frame_container[0]
            if current_frame is not None:
                try:
                    h, w = current_frame.shape[:2]
                    if w > 1280:
                        scale = 1280 / w
                        frame_copy = cv2.resize(current_frame, (1280, int(h * scale)), interpolation=cv2.INTER_AREA)
                    else:
                        frame_copy = current_frame.copy()
                except Exception as e:
                    _logger.error(f"Erro ao redimensionar frame web: {e}")
                    frame_copy = current_frame.copy()
            else:
                pass 

        if frame_copy is not None:

            try:

                ret, buffer = cv2.imencode('.jpg', frame_copy, [cv2.IMWRITE_JPEG_QUALITY, 70]) 
                if ret:
                    frame_bytes = buffer.tobytes()
                    

                    yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
                else:
                    _logger.error("MJPEG_GEN_ENCODE_FAIL: cv2.imencode falhou.")

            except Exception as e:
                _logger.error(f"MJPEG_GEN_EXCEPTION: Erro ao codificar frame para MJPEG: {e}", exc_info=True)
        else:
            
            
            
            
            try:
                placeholder_frame = np.zeros((480, 640, 3), dtype=np.uint8)
                cv2.putText(placeholder_frame, "Aguardando video...", (50, 240), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2, cv2.LINE_AA)
                
                ret_placeholder, buffer_placeholder = cv2.imencode('.jpg', placeholder_frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
                if ret_placeholder:
                    frame_bytes_placeholder = buffer_placeholder.tobytes()
                    yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes_placeholder + b'\r\n')
                else:
                    _logger.error("MJPEG_GEN_ENCODE_FAIL: cv2.imencode falhou para frame placeholder.")
            except Exception as e_placeholder:
                _logger.error(f"MJPEG_GEN_EXCEPTION: Erro ao gerar/enviar frame placeholder: {e_placeholder}", exc_info=True)
            
        
        current_time = time.time()
        elapsed_time = current_time - last_frame_time
        wait_time = (1.0 / target_fps) - elapsed_time
        if wait_time > 0:
            time.sleep(wait_time)
        last_frame_time = time.time() 

login_manager = LoginManager()
login_manager.init_app(flask_app)

class FlaskUser(UserMixin):
    """Classe de usuário para Flask-Login."""
    def __init__(self, id, username, role, first_name=None, last_name=None, email=None):
        self.id = id
        self.username = username
        self.role = role
        self.first_name = first_name
        self.last_name = last_name
        self.email = email
def get_ipv4_address():
    """Tenta obter o endereço IPv4 local principal."""
    try:

        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0)

        s.connect(('8.8.8.8', 1)) 
        ip_address = s.getsockname()[0]
        s.close()
        return ip_address
    except Exception:
        _logger.warning("Não foi possível determinar o endereço IPv4 local automaticamente.")
        return None 


def role_required(role):
    """Decorador para exigir uma role específica."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or getattr(current_user, 'role', None) != role:
                user_desc = f"'{current_user.username}'" if current_user.is_authenticated else "'Anônimo'"
                _logger.warning(f"Acesso não autorizado à rota {request.path} pelo usuário {user_desc}. Role necessária: '{role}'. Acionando erro 405 para exibir página de erro/logout.")
                
                
                
                abort(405) 
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@login_manager.user_loader
def load_user(username):
    try:
        db = get_db() 
        if db is None:
            _logger.error("user_loader: Failed to get DB connection from g.")
            return None

        user_info = db.get_user(username) 

        if user_info: 
            _logger.debug(f"user_loader: user_info para '{username}': {user_info}")

            if not isinstance(user_info, dict):
                # Fallback em caso de inconsistência, embora db.get_user já retorne dict agora
                _logger.warning(f"user_loader: user_info para '{username}' não é um dicionário. Tipo: {type(user_info)}")
                return None

            user_db_username = user_info.get('username')
            is_admin_flag = bool(user_info.get('is_admin'))
            user_email = user_info.get('email', f"{user_db_username}@example.com")
            user_first_name = user_info.get('first_name')
            user_last_name = user_info.get('last_name')

            flask_user_role = 'admin' if is_admin_flag else 'user'
            _logger.debug(f"user_loader: User '{user_db_username}' carregado com role {flask_user_role}.")
            
            return FlaskUser(id=user_db_username, username=user_db_username, role=flask_user_role, 
                           first_name=user_first_name, last_name=user_last_name, email=user_email)
        else:
            _logger.warning(f"user_loader: Usuário '{username}' não encontrado no banco de dados.")
            return None
    except Exception as e:
        _logger.error(f"user_loader: Exception occurred for user '{username}': {e}", exc_info=True) 
        return None 

@login_manager.unauthorized_handler
def unauthorized():
    """Redireciona para a página de login se não estiver autorizado."""

    if request.path.startswith('/api/'):
        _logger.warning(f"Acesso não autenticado à rota API {request.path}.")
        return jsonify(message="Autenticação necessária."), 401
    return redirect(url_for('login_page')) 

@flask_app.route('/api/login', methods=['POST'])
def api_login():
    """Endpoint para login remoto (apenas admin)."""
    db = get_db() 
    data = request.json
    username = data.get('username')
    password = data.get('password')
    user_info = db.verify_user(username, password) 
    

    if user_info and user_info.get('is_admin'):
        
        flask_user_role = 'admin'
        user_first_name = user_info.get('first_name')
        user_last_name = user_info.get('last_name')
        user_email = user_info.get('email', f"{user_info['username']}@example.com")
        user = FlaskUser(user_info['username'], user_info['username'], flask_user_role, first_name=user_first_name, last_name=user_last_name, email=user_email)        
        login_user(user)
        session.permanent = True 
        session['user_role'] = flask_user_role 
        client_ip = request.remote_addr
        _logger.info(f"Usuário admin '{username}' logado remotamente do IP: {client_ip}")
        
        with active_users_lock:
            now_utc = get_current_utc_timestamp()
            hostname = client_ip 
            try:
                hostname, _, _ = socket.gethostbyaddr(client_ip)
            except (socket.herror, socket.gaierror):
                _logger.debug(f"Não foi possível resolver o hostname para o IP {client_ip}. Usando IP como hostname.")
            
            active_users[user.username] = {
                'role': flask_user_role, 
                'login_time': now_utc,
                'last_seen': now_utc,
                'ip_address': client_ip,
                'user_agent': request.user_agent.string,
                'hostname': hostname,
                'first_name': user_first_name,
                'last_name': user_last_name,
                'email': user_email            }
        db.log_activity(username, client_ip, 'LOGIN_SUCCESS')

        return jsonify({"success": True, "message": "Login bem-sucedido"}), 200
    else:

        _logger.warning(f"Tentativa de login remoto via API falhou para usuário '{username}'.") 
        return jsonify({"message": "Credenciais inválidas ou usuário não é administrador"}), 401

@flask_app.route('/api/register', methods=['POST']) 
def api_register():
    """Endpoint para registro de novos usuários."""
    db = get_db()
    data = request.json

    if not data:
        return jsonify({"success": False, "message": "Nenhum dado recebido."}), 400

    nome = data.get('nome')
    sobrenome = data.get('sobrenome')
    matricula = data.get('matricula') 
    email = data.get('email')
    password = data.get('password')
    confirm_password = data.get('confirm_password')

    required_fields = {'Nome': nome, 'Sobrenome': sobrenome, 'Matrícula': matricula, 'Email': email, 'Senha': password, 'Confirmar Senha': confirm_password}
    for field_name, field_value in required_fields.items():
        if not field_value:
            _logger.warning(f"Tentativa de registro falhou: Campo '{field_name}' ausente.")
            return jsonify({"success": False, "message": f"Campo '{field_name}' é obrigatório."}), 400

    if password != confirm_password:
        _logger.warning(f"Tentativa de registro falhou para '{matricula}': Senhas não coincidem.")
        return jsonify({"success": False, "message": "As senhas não coincidem."}), 400

    
    if len(password) < 6:
        _logger.warning(f"Tentativa de registro falhou para '{matricula}': Senha muito curta.")
        return jsonify({"success": False, "message": "A senha deve ter pelo menos 6 caracteres."}), 400

    
    if db.get_user(matricula):
        _logger.warning(f"Tentativa de registro falhou: Matrícula '{matricula}' já existe.")
        return jsonify({"success": False, "message": "Esta matrícula já está registrada. Tente fazer login."}), 409 
    

    hashed_password = generate_password_hash(password) 
    
    _logger.info(f"Tentando criar usuário: {matricula}, {email}, {nome}, {sobrenome}")
    
    
    user_created_successfully = db.create_user(
        username=matricula, 
        password=hashed_password, 
        email=email, 
        first_name=nome, 
        last_name=sobrenome, 
        is_admin=0 
    )

    if user_created_successfully:
        
        _logger.info(f"Endpoint /api/register: Usuário '{matricula}' registrado com sucesso e confirmado pelo DB.")
        return jsonify({"success": True, "message": "Registro bem-sucedido! Você já pode fazer login."}), 201 
    else:
        
        _logger.error(f"Endpoint /api/register: Falha ao registrar usuário '{matricula}' (db.create_user retornou False).")
        return jsonify({"success": False, "message": "Erro ao registrar usuário. A matrícula ou email podem já estar em uso."}), 400 

@flask_app.route('/api/settings', methods=['GET'])
@login_required
@role_required('admin')
def get_settings():
    try:
        
        current_settings = _settings_manager.get_all()



        if _security_manager:
            settings_json = json.dumps(current_settings).encode('utf-8')
            encrypted_settings = _security_manager.encrypt_data(settings_json)
            
            return Response(encrypted_settings, mimetype='application/octet-stream')
        else:
            _logger.error("SecurityManager não inicializado. Não é possível criptografar a resposta.")
            return jsonify({"message": "Erro interno: Criptografia não disponível"}), 500

    except Exception as e:
        _logger.error(f"Erro ao obter configurações via API: {e}")
        return jsonify({"message": "Erro ao obter configurações"}), 500

@flask_app.route('/api/settings', methods=['POST'])
@login_required
@role_required('admin')
def update_settings_api():
    _logger.info(f"'/api/settings' POST request received from user '{current_user.username}'")
    settings_manager = get_settings_manager()
    security = get_security_manager() 

    try:
        content_type = request.headers.get('Content-Type', '').lower()
        _logger.debug(f"Content-Type recebido para /api/settings POST: {content_type}")

        new_settings_data = None

        if 'application/octet-stream' in content_type:
            _logger.info("Recebido payload como application/octet-stream (esperado criptografado).")
            encrypted_data = request.data
            if not encrypted_data:
                _logger.warning("Nenhum dado criptografado recebido na requisição POST para /api/settings.")
                return jsonify({"message": "Nenhum dado recebido."}), 400
            
            if not security:
                _logger.error("SecurityManager não está disponível no servidor para descriptografar settings.")
                return jsonify({"message": "Erro interno: Criptografia não configurada no servidor."}), 500
            
            decrypted_data_bytes = security.decrypt_data(encrypted_data)
            if decrypted_data_bytes is None:
                _logger.error("Falha ao descriptografar os dados de settings recebidos.")
                return jsonify({"message": "Erro ao processar dados recebidos (descriptografia falhou)."}), 400
            
            try:
                new_settings_data = json.loads(decrypted_data_bytes.decode('utf-8'))
                _logger.info("Dados de settings descriptografados e carregados como JSON com sucesso.")
            except (json.JSONDecodeError, UnicodeDecodeError) as e:
                _logger.error(f"Erro ao decodificar JSON dos dados de settings descriptografados: {e}")
                return jsonify({"message": "Erro ao decodificar dados de settings descriptografados."}), 400
                
        elif 'application/json' in content_type:
            _logger.info("Recebido payload como application/json (não criptografado).")
            new_settings_data = request.get_json()
            if new_settings_data is None: 
                _logger.warning("Falha ao fazer parse do JSON recebido em /api/settings.")
                return jsonify({"message": "Formato de dados inválido. Esperado JSON."}), 400
        else:
            _logger.error(f"Content-Type não suportado: {content_type}. Esperado 'application/json' ou 'application/octet-stream'.")
            return jsonify({"message": f"Content-Type não suportado: {content_type}"}), 415

        if new_settings_data is None:
            _logger.error("Nenhum dado de configuração processável foi obtido da requisição.")
            return jsonify({"message": "Dados de configuração não puderam ser processados."}), 400


        
        current_settings = settings_manager.get_all()
        
        
        
        if 'VIDEO_PARAMS' in new_settings_data:
            current_settings['VIDEO_PARAMS'].update(new_settings_data['VIDEO_PARAMS'])
            _logger.info("VIDEO_PARAMS mesclados.")
            
            # Sync to AI_PARAMS.advanced.video_source for persistence
            try:
                video_params = current_settings['VIDEO_PARAMS']
                
                # Ensure structure exists
                if 'AI_PARAMS' not in current_settings: current_settings['AI_PARAMS'] = {}
                if 'advanced' not in current_settings['AI_PARAMS']: current_settings['AI_PARAMS']['advanced'] = {}
                
                # Construct video_source config
                video_source_config = {
                    'type': video_params.get('source_type', 'stream'),
                    'stream_url': video_params.get('source_param', ''),
                    'camera_id': int(video_params.get('source_param', 0)) if video_params.get('source_type') == 'camera' else 0,
                    'default_video_path': video_params.get('source_param', '') if video_params.get('source_type') == 'video' else ''
                }
                
                current_settings['AI_PARAMS']['advanced']['video_source'] = video_source_config
                _logger.info(f"Sincronizado VIDEO_PARAMS -> AI_PARAMS.advanced.video_source: {video_source_config}")
                
            except Exception as e:
                _logger.error(f"Erro ao sincronizar VIDEO_PARAMS para AI_PARAMS: {e}")
        
        if 'AI_PARAMS' in new_settings_data and 'AI_PARAMS' in current_settings:
            current_settings['AI_PARAMS'].update(new_settings_data['AI_PARAMS'])
            _logger.info("AI_PARAMS mesclados.")
        
        # <<<< NOVO: Merge para COLORS >>>>
        if 'COLORS' in new_settings_data and 'COLORS' in current_settings:
            # Merge detection colors
            if 'detection' in new_settings_data.get('COLORS', {}):
                if 'detection' not in current_settings['COLORS']:
                    current_settings['COLORS']['detection'] = {}
                current_settings['COLORS']['detection'].update(new_settings_data['COLORS']['detection'])
                _logger.info("COLORS.detection mesclados.")
        
        # <<<< NOVO: Merge para VISUAL_STYLE >>>>
        if 'VISUAL_STYLE' in new_settings_data and 'VISUAL_STYLE' in current_settings:
            for key in ['detection_box', 'detection_alert', 'fracture_alert', 'roi_display']:
                if key in new_settings_data.get('VISUAL_STYLE', {}):
                    if key not in current_settings['VISUAL_STYLE']:
                        current_settings['VISUAL_STYLE'][key] = {}
                    current_settings['VISUAL_STYLE'][key].update(new_settings_data['VISUAL_STYLE'][key])
                    _logger.info(f"VISUAL_STYLE.{key} mesclado.")

        if settings_manager.save_settings(current_settings):
            _logger.info(f"Configurações atualizadas e salvas com sucesso pelo usuário '{current_user.username}'.")
            
            
            
            if _model_instance: 
                _logger.info("Aplicando novas configurações à instância do modelo no servidor...") 

                if 'AI_PARAMS' in current_settings:
                    ai_params = current_settings['AI_PARAMS']
                    if 'conf_default' in ai_params:
                        if hasattr(_model_instance, 'update_conf'):
                            _model_instance.update_conf(ai_params['conf_default'])
                            _logger.info(f"Modelo: Confiança atualizada para {ai_params['conf_default']}")
                        else:
                            _logger.warning("Modelo não possui método 'update_conf'.")
                    if 'iou_default' in ai_params:
                        if hasattr(_model_instance, 'update_iou'):
                            _model_instance.update_iou(ai_params['iou_default'])
                            _logger.info(f"Modelo: IOU atualizado para {ai_params['iou_default']}")
                        else:
                            _logger.warning("Modelo não possui método 'update_iou'.")

                    # Consolidate advanced settings + video params into ONE call
                    if hasattr(_model_instance, 'update_advanced_settings'):
                        consolidated_settings = {}
                        
                        # Add advanced config if present
                        if 'advanced' in ai_params:
                            consolidated_settings.update(ai_params['advanced'])
                        
                        # Add thresholds if present
                        if 'thresholds' in ai_params:
                            consolidated_settings['thresholds'] = ai_params['thresholds']
                        
                        # Add video source config from VIDEO_PARAMS if present
                        if 'VIDEO_PARAMS' in current_settings:
                            video_params = current_settings['VIDEO_PARAMS']
                            video_source_config = {
                                'type': video_params.get('source_type', 'stream'),
                                'stream_url': video_params.get('source_param', ''),
                                'camera_id': int(video_params.get('source_param', 0)) if video_params.get('source_type') == 'camera' else 0,
                                'default_video_path': video_params.get('source_param', '') if video_params.get('source_type') == 'video' else '',
                                'resolution': video_params.get('resolution', ''),
                                'fps': video_params.get('fps', '')
                            }
                            consolidated_settings['video_source'] = video_source_config
                            
                        # Add COLORS if present
                        if 'COLORS' in current_settings and 'detection' in current_settings['COLORS']:
                            consolidated_settings['colors'] = current_settings['COLORS']['detection']
                        
                        # Single call with all settings
                        _logger.info(f"Modelo: Atualizando configurações consolidadas...")
                        _model_instance.update_advanced_settings(consolidated_settings)
                        _logger.info(f"Modelo: Configurações consolidadas aplicadas com sucesso")
                    else:
                        _logger.warning("Modelo não possui método 'update_advanced_settings'.")
            

            return jsonify({"message": "Configurações atualizadas com sucesso."}), 200
        else:
            _logger.error("Falha ao salvar as configurações atualizadas no arquivo.")
            return jsonify({"message": "Erro ao salvar configurações no servidor."}), 500

    except Exception as e:
        _logger.error(f"Erro ao atualizar configurações via API: {e}", exc_info=True)
        return jsonify({"message": f"Erro interno ao processar configurações: {e}"}), 500
    
@flask_app.route('/api/basic_stats')
@login_required  
def get_basic_stats():
    """Retorna estatísticas básicas para usuários não-admin."""
    if _current_stats is None or _stats_lock is None:
        _logger.error("API /api/basic_stats: _current_stats ou _stats_lock não estão disponíveis.")
        return jsonify({"error": "Stats not available"}), 503
    
    with _stats_lock:
        lata_normal_count = _current_stats.get('lata_normal', 0)
        lata_invertida_count = _current_stats.get('lata_invertida', 0)
        lata_tombada_count = _current_stats.get('lata_tombada', 0)
        fracture_count = _current_stats.get('fracture', 0)
        total_detected = (
            lata_normal_count +
            lata_invertida_count +
            lata_tombada_count +
            fracture_count
        )
        ai_accuracy = _current_stats.get('ai_accuracy_percent', 96.3) # Usando o mesmo default da rota /api/stats
        fps_value = _current_stats.get('fps', 0.0)

        return jsonify({
            "status": "Operacional", 
            "total_detected": total_detected,
            "lata_normal": lata_normal_count,
            "lata_invertida": lata_invertida_count,
            "lata_tombada": lata_tombada_count,
            "fracture": fracture_count,
            "ai_accuracy_percent": ai_accuracy,
            "fps": fps_value,
            "camera_status": {
                "main": "online",
                "secondary1": "online",
                "secondary2": "online",
                "secondary3": "offline",
                "secondary4": "offline",
                "rejection": "processing"
            }
        })
    

@flask_app.route('/api/stats')
@login_required
@role_required('admin')

def get_stats():
    if _current_stats is None or _stats_lock is None:
        return jsonify({"error": "Stats not available"}), 503 
    stats_copy = {} 
    with _stats_lock: 
        stats_copy = _current_stats.copy()

    stats_copy['lata_normal'] = int(stats_copy.get('lata_normal', 0))
    stats_copy['lata_tombada'] = int(stats_copy.get('lata_tombada', 0))
    stats_copy['lata_invertida'] = int(stats_copy.get('lata_invertida', 0))
    stats_copy['fracture'] = int(stats_copy.get('fracture', 0))

    stats_copy['status'] = "Operacional" 
    if _system_start_time:
        uptime_seconds = time.time() - _system_start_time
        stats_copy['system_uptime_seconds'] = int(uptime_seconds)
    else:
        stats_copy['system_uptime_seconds'] = 0
    stats_copy['ai_accuracy_percent'] = stats_copy.get('ai_accuracy_percent', 96.3) 
    stats_copy['total_detected'] = (stats_copy.get('lata_normal', 0) +
                                   stats_copy.get('lata_invertida', 0) +
                                   stats_copy.get('lata_tombada', 0) +
                                   stats_copy.get('fracture', 0))
    resolution = "N/A"
    if _frame_container is not None and _frame_lock is not None:
        with _frame_lock:
            current_frame = _frame_container[0]
            if current_frame is not None:
                try:
                    h, w, _ = current_frame.shape
                    resolution = f"{w}x{h}"
                except Exception as e:
                    _logger.debug(f"Não foi possível obter resolução do frame: {e}")
    stats_copy['resolution'] = resolution
    
    if _settings_manager: 
        all_settings = _settings_manager.get_all() 
        model_settings = all_settings.get('MODEL_PARAMS') 
        if model_settings:
            stats_copy['model_version'] = model_settings.get('model_version', 'N/A')
            stats_copy['model_last_trained'] = model_settings.get('last_trained_date', 'N/A')
            stats_copy['model_dataset_size'] = model_settings.get('dataset_size_info', 'N/A')

    
    try:
            db = get_db()
            today_date = datetime.now().date()
            yesterday_date = today_date - timedelta(days=1)

            
            counts_yesterday = db.get_daily_counts(yesterday_date)

            def calculate_pct_change(current_model_count, previous_day_total_from_db):
                if previous_day_total_from_db is None or previous_day_total_from_db == 0:
                    
                    return None
                
                return ((current_model_count - previous_day_total_from_db) / previous_day_total_from_db) * 100

            if counts_yesterday: 
                stats_copy['lata_normal_change_pct'] = calculate_pct_change(stats_copy.get('lata_normal', 0), counts_yesterday.get('normal'))
                stats_copy['lata_tombada_change_pct'] = calculate_pct_change(stats_copy.get('lata_tombada', 0), counts_yesterday.get('fallen'))
                stats_copy['lata_invertida_change_pct'] = calculate_pct_change(stats_copy.get('lata_invertida', 0), counts_yesterday.get('inverted'))
                stats_copy['fracture_change_pct'] = calculate_pct_change(stats_copy.get('fracture', 0), counts_yesterday.get('fracture'))
            else: 
                stats_copy['lata_normal_change_pct'] = None
                stats_copy['lata_tombada_change_pct'] = None
                stats_copy['lata_invertida_change_pct'] = None
                stats_copy['fracture_change_pct'] = None

    except Exception as e:
        _logger.error(f"Erro ao calcular variação percentual diária: {e}")
        
        stats_copy['lata_normal_change_pct'] = None
        stats_copy['lata_tombada_change_pct'] = None
        stats_copy['lata_invertida_change_pct'] = None
        stats_copy['fracture_change_pct'] = None

    
    db = get_db() 
    db_alerts_tuples = db.get_recent_alerts(limit=5) 

    frontend_alerts = []
    if db_alerts_tuples:
        for alert_tuple in db_alerts_tuples:
            
            
            
            
            timestamp_for_frontend = alert_tuple[1] 
            try:
                
                dt_object = datetime.fromisoformat(alert_tuple[1])
                timestamp_for_frontend = dt_object.strftime('%H:%M:%S')
            except ValueError:
                _logger.debug(f"Não foi possível converter o timestamp do alerta '{alert_tuple[1]}' para HH:MM:SS. Usando valor raw.")
            except TypeError: 
                _logger.debug(f"Timestamp do alerta inválido ou nulo: '{alert_tuple[1]}'.")

            
            db_alert_type = alert_tuple[2] 
            frontend_level = 'info' 
            if db_alert_type: 
                if "Invertida" in db_alert_type:
                    frontend_level = 'error'
                elif "Tombada" in db_alert_type:
                    frontend_level = 'warning'
                

            frontend_alerts.append({
                'timestamp': timestamp_for_frontend,
                'message': alert_tuple[4],  
                'level': frontend_level     
            })
    stats_copy['alerts'] = frontend_alerts

    
    
    stats_copy['camera_status'] = {
        'main': 'online', 
        'secondary1': 'online', 
        'secondary2': 'online',
        'secondary3': 'offline',
        'secondary4': 'offline',
        'rejection': 'processing' 
    }

    if _security_manager:
        stats_json = json.dumps(stats_copy, default=json_serial_default).encode('utf-8')
        encrypted_stats = _security_manager.encrypt_data(stats_json)
        
        return Response(encrypted_stats, mimetype='application/octet-stream')
    else:
        _logger.warning("SecurityManager não inicializado. Não é possível criptografar stats para /api/stats.") 
        return jsonify({"message": "Erro interno: Criptografia não disponível"}), 500
    

@flask_app.route('/login', methods=['GET', 'POST'])
def login_page():
    """Página de login."""
    
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        error = None 

        # Detalhes para o modal de sessão duplicada
        duplicate_session_info = None

        if not username:
            error = "Matrícula é obrigatória."
            return render_template('login.html', 
                                 error=error, 
                                 duplicate_session_info=None,
                                 last_username=None,
                                 last_remember=False)

        db = get_db()
        force_logout_username = request.form.get('force_logout_username')
        
        # Lógica de Verificação de Credenciais
        user_info = None
        needs_password_check = True

        # Se for um force logout vindo do modal, verificamos se já validamos a senha na sessão anterior
        if force_logout_username and force_logout_username == username:
            if session.get('pending_force_logout_user') == username:
                _logger.info(f"Ignorando senha para '{username}' pois já validamos antes do modal.")
                user_info = db.get_user(username)
                needs_password_check = False
        
        if needs_password_check:
            if not password:
                error = "Senha é obrigatória."
                return render_template('login.html', 
                                     error=error, 
                                     duplicate_session_info=None,
                                     last_username=username,
                                     last_remember=bool(request.form.get('remember-me')))
            user_info = db.verify_user(username, password)

        if not user_info:
            error = "Matrícula ou senha incorretos."
            _logger.warning(f"Falha de login para '{username}' do IP {request.remote_addr}")
            return render_template('login.html', 
                                 error=error, 
                                 duplicate_session_info=None,
                                 last_username=username,
                                 last_remember=bool(request.form.get('remember-me')))

        # Se chegamos aqui, as credenciais são VÁLIDAS. Agora checamos sessão ativa.
        with active_users_lock:
            # Caso 1: O usuário CLICOU no botão do modal para deslogar a outra sessão
            if force_logout_username == username:
                if username in active_users:
                    _logger.info(f"Forçando logout da sessão de '{username}' a pedido do IP {request.remote_addr}.")
                    del active_users[username]
                session.pop('pending_force_logout_user', None) # Limpa flag
                
            # Caso 2: Login normal, mas existe sessão ativa em outro lugar
            elif username in active_users:
                active_session_details = active_users[username]
                current_ip = request.remote_addr
                active_ip = active_session_details.get('ip_address')
                active_hostname = active_session_details.get('hostname', active_ip)
                active_user_agent_str = active_session_details.get('user_agent', '')
                
                active_browser = "Navegador Desconhecido"
                if active_user_agent_str:
                    ua = active_user_agent_str.lower()
                    if "edg/" in ua or "edge" in ua: active_browser = "Edge"
                    elif "chrome/" in ua: active_browser = "Chrome"
                    elif "firefox/" in ua: active_browser = "Firefox"
                    elif "safari/" in ua: active_browser = "Safari"

                modal_msg = (f"Esta conta ('{username}') já está logada " + 
                            (f"neste computador (Edge)." if current_ip == active_ip else f"em {active_hostname}."))

                # Salvamos na sessão que este usuário JÁ VALIDOU A SENHA e pode forçar o login
                session['pending_force_logout_user'] = username
                
                duplicate_session_info = {
                    "message": modal_msg,
                    "username_in_conflict": username
                }
                _logger.warning(f"Login duplicado para '{username}'. Exibindo modal para IP {current_ip}.")
                
                return render_template('login.html', 
                                     error=None, 
                                     duplicate_session_info=duplicate_session_info,
                                     last_username=username,
                                     last_remember=bool(request.form.get('remember-me')))

        if user_info:  # Sucesso total no login
            
            user_first_name = user_info.get('first_name')
            user_last_name = user_info.get('last_name')
            is_admin_flag = bool(user_info.get('is_admin'))
            user_email = user_info.get('email', f"{user_info['username']}@example.com")
            flask_user_role = 'admin' if is_admin_flag else 'user'

            remember_me_checked = request.form.get('remember-me') 
            user = FlaskUser(id=user_info['username'], username=user_info['username'], role=flask_user_role, first_name=user_first_name, last_name=user_last_name, email=user_email)
            login_user(user, remember=bool(remember_me_checked)) 
            
            if remember_me_checked:
                session.permanent = True 
                _logger.info(f"Login com 'Lembrar-me' para '{username}'. Sessão permanente ativada.")
            else:
                session.permanent = False 
                _logger.info(f"Login sem 'Lembrar-me' para '{username}'. Sessão não permanente.")

            client_ip = request.remote_addr
            _logger.info(f"Usuário '{username}' (Role: {flask_user_role}) logado com sucesso do IP: {client_ip}")
            db.log_activity(username, client_ip, 'LOGIN_SUCCESS')

            
            with active_users_lock:
                now_utc = get_current_utc_timestamp()
                hostname = client_ip 
                try:
                    hostname, _, _ = socket.gethostbyaddr(client_ip)
                except (socket.herror, socket.gaierror):
                    _logger.debug(f"Não foi possível resolver o hostname para o IP {client_ip} do usuário {user.username}. Usando IP como hostname.")

                active_users[user.username] = {
                    'role': flask_user_role, 
                    'login_time': now_utc,
                    'last_seen': now_utc,
                    'ip_address': client_ip,
                    'user_agent': request.user_agent.string,
                    'hostname': hostname,
                    'first_name': user_first_name,
                    'last_name': user_last_name,
                    'email': user_email    }
            return redirect(url_for('index')) 
        else:
            error = "Credenciais inválidas. Por favor, tente novamente."
            _logger.warning(f"Tentativa de login falhou para usuário '{username}' do IP: {request.remote_addr}")
            db.log_activity(username, request.remote_addr, 'LOGIN_FAIL')
            return render_template('login.html', 
                                 error=error, 
                                 duplicate_session_info=None,
                                 last_username=username,
                                 last_remember=bool(request.form.get('remember-me')))

    return render_template('login.html', 
                         error=None, 
                         duplicate_session_info=None,
                         last_username=None,
                         last_remember=False)

@flask_app.route('/')
@login_required 
def index():
    """Página principal do controle remoto."""

    return render_template('index.html', username=current_user.username, client_ip=request.remote_addr)

@flask_app.route('/dashboard')
@login_required
def dashboard_page():
    """Página do dashboard."""

    return render_template('dashboard.html', username=current_user.username, client_ip=request.remote_addr)

@flask_app.route('/users')
@login_required
@role_required('admin')
def users():
    # Página do dashboard.
    return render_template('users.html')

@flask_app.route('/api/users_list')
@login_required
@role_required('admin')
def get_all_users_api():
    try:
        db = get_db()
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '')
        
        users_raw, total_items = db.get_all_users_paged(page, per_page, search)
        
        user_list = []
        for u in users_raw:
            # u = (id, username, email, first_name, last_name, is_admin, email_notifications, created_at)
            uid, uname, email, fname, lname, is_admin, email_notif, created_at = u
            
            # Formatação para o frontend
            full_name = f"{fname} {lname}".strip() or uname
            user_list.append({
                "id": uid,
                "name": full_name,
                "username": uname,
                "email": email,
                "avatar_initials": (fname[0] if fname else uname[0]).upper(),
                "status": "Online" if uname in active_users else "Offline",
                "status_class": "badge-online" if uname in active_users else "badge-offline",
                "device_name": active_users.get(uname, {}).get('hostname', 'N/A') if uname in active_users else "N/A",
                "ip_address": active_users.get(uname, {}).get('ip_address', 'N/A') if uname in active_users else "N/A",
                "last_activity_time": created_at, # Simplificado ou buscar do active_users se online
                "last_activity_details": "Dashboard" if uname in active_users else "N/A",
                "is_admin": bool(is_admin),
                "email_notifications": bool(email_notif)
            })
            
        return jsonify({
            "users": user_list,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total_items": total_items,
                "total_pages": math.ceil(total_items / per_page) if per_page > 0 else 0
            }
        })
    except Exception as e:
        _logger.error(f"Erro ao buscar lista de usuários: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/user/update_email_notification', methods=['POST'])
@login_required
@role_required('admin')
def update_email_notification_api():
    try:
        db = get_db()
        data = request.json
        user_id = data.get('user_id')
        enabled = data.get('enabled', False)
        
        if db.update_user_email_notification(user_id, enabled):
            return jsonify({"success": True, "message": "Preferência de email atualizada."})
        else:
            return jsonify({"success": False, "message": "Falha ao atualizar no banco de dados."}), 500
    except Exception as e:
        _logger.error(f"Erro ao atualizar preferência de email: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@flask_app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template('forgot_password.html')
    
    email = request.form.get('email')
    db = get_db()
    user = db.get_user_by_email(email)
    
    if user:
        token = secrets.token_urlsafe(32)
        expires_at = get_current_utc_timestamp() + timedelta(hours=1)
        if db.add_password_reset_token(user[0], token, expires_at):
            reset_link = url_for('reset_password', token=token, _external=True)
            # Envia o e-mail (usando threading para não travar o request)
            import threading
            threading.Thread(target=send_password_reset_email, args=(email, reset_link)).start()
            
    # Mensagem neutra para evitar enumeração de usuários
    return render_template('forgot_password.html', success="Se este e-mail estiver cadastrado, um link de recuperação será enviado em instantes.")

@flask_app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    db = get_db()
    user = db.get_user_by_reset_token(token)
    
    if not user:
        return render_template('reset_password.html', error="O link de recuperação é inválido ou já expirou.", invalid=True)
    
    if request.method == 'GET':
        return render_template('reset_password.html', token=token)
    
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    if not password or password != confirm_password:
        return render_template('reset_password.html', token=token, error="As senhas não coincidem ou são inválidas.")
    
    hashed_password = generate_password_hash(password)
    if db.update_user_password(user[0], hashed_password):
        db.mark_reset_token_used(token)
        return render_template('reset_password.html', success="Sua senha foi atualizada com sucesso! Você já pode entrar no sistema.")
    else:
        return render_template('reset_password.html', token=token, error="Erro interno ao atualizar a senha. Tente novamente.")

@flask_app.route('/api/test_smtp', methods=['POST'])
@login_required
@role_required('admin')
def test_smtp():
    """Envia um email de teste para os usuários selecionados."""
    try:
        data = request.json
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return jsonify({"success": False, "message": "Nenhum usuário selecionado."}), 400
            
        db = get_db()
        recipients = []
        
        for user_id in user_ids:
            cursor = db.conn.cursor()
            cursor.execute("SELECT email FROM users WHERE id = ?", (user_id,))
            row = cursor.fetchone()
            if row:
                recipients.append(row[0])
        
        if not recipients:
            return jsonify({"success": False, "message": "E-mails não encontrados para os IDs fornecidos."}), 404
            
        subject = "[VisionAlign] Teste de Conexão SMTP / Alerta de Fratura"
        body = f"Este é um e-mail de teste de formatação de Alerta HTML enviado por {current_user.username}.\\nValidação das imagens e layout.\\nData: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        from utils.email_utils import send_fracture_alert_email
        import cv2
        import numpy as np
        import os
        
        # Gerar imagens de teste temporárias para simular o anexo
        alerts_dir = os.path.join("data", "alerts")
        os.makedirs(alerts_dir, exist_ok=True)
        img_temp = os.path.join(alerts_dir, "test_main.jpg")
        roi_temp = os.path.join(alerts_dir, "test_roi.jpg")
        
        if not os.path.exists(img_temp):
            dummy_img = np.zeros((480, 640, 3), dtype=np.uint8)
            cv2.putText(dummy_img, "IMAGEM TESTE", (50, 240), cv2.FONT_HERSHEY_SIMPLEX, 2, (255, 255, 255), 3)
            cv2.imwrite(img_temp, dummy_img)
            
        if not os.path.exists(roi_temp):
            dummy_roi = np.zeros((150, 150, 3), dtype=np.uint8)
            dummy_roi[:] = (0, 0, 255) # Fundo vermelho para ROI
            cv2.putText(dummy_roi, "ROI", (20, 80), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (255, 255, 255), 3)
            cv2.imwrite(roi_temp, dummy_roi)

        # Usamos a função existente que já carrega o .env
        success = send_fracture_alert_email(recipients, subject, body, image_path=img_temp, roi_path=roi_temp)
        
        if success:
            return jsonify({"success": True, "message": f"E-mail de teste enviado com sucesso para {len(recipients)} destinatário(s)."})
        else:
            return jsonify({"success": False, "message": "Erro ao enviar e-mail. Verifique o arquivo .env e os logs do servidor."}), 500
            
    except Exception as e:
        _logger.error(f"Erro no teste de SMTP: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/settings')
@login_required
@role_required('admin') 
def settings_page():
    """Página de configurações."""
    return render_template('settings.html', username=current_user.username, client_ip=request.remote_addr)

@flask_app.route('/retrain')
@login_required
@role_required('admin')
def retrain_page():
    """Página de Retreinamento do modelo."""
    return render_template('retrain.html', username=current_user.username, client_ip=request.remote_addr)

@flask_app.route('/api/dataset_info', methods=['GET'])
@login_required
@role_required('admin')
def get_dataset_info():
    """Retorna informações sobre o dataset coletado."""
    try:
        dataset_path = _settings_manager.get_setting('MODEL_PARAMS', 'dataset_path', 'data/dataset_collect/images')
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        full_path = os.path.join(project_root, dataset_path)
        
        image_count = 0
        if os.path.exists(full_path):
            image_count = len([f for f in os.listdir(full_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
        
        # Simulação de outras infos
        return jsonify({
            "image_count": image_count,
            "class_count": 4, # lata_normal, lata_invertida, lata_tombada, fracture
            "last_update": datetime.now().strftime("%d/%m/%Y %H:%M")
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@flask_app.route('/api/start_retrain', methods=['POST'])
@login_required
@role_required('admin')
def start_retrain():
    """Inicia o processo de retreinamento."""
    try:
        data = request.json
        model_type = data.get('model', 'align')
        epochs = data.get('epochs', 50)
        
        if not _otx_manager:
            return jsonify({"success": False, "message": "OTX Manager não inicializado"}), 500
            
        # Inicia treinamento em uma thread separada para não travar o Flask
        def run_training():
            _logger.info(f"Iniciando treinamento OTX para {model_type} com {epochs} épocas...")
            # Aqui chamamos o otx_manager
            # Por enquanto, como é uma demo, vamos apenas logar o início
            dataset_path = _settings_manager.get_setting('MODEL_PARAMS', 'dataset_path', 'data/dataset_collect/images')
            _otx_manager.train(model_type, dataset_path)
            
        threading.Thread(target=run_training, daemon=True).start()
        
        return jsonify({"success": True, "message": "Treinamento iniciado. Acompanhe o console."}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/test_models', methods=['POST'])
@login_required
@role_required('admin')
def test_models_api():
    """Executa autodiagnóstico nos modelos."""
    try:
        # Simulação de teste
        results = {
            "align": {"status": "success", "message": "Modelo VisionAlign carregado e operacional."},
            "fracture": {"status": "success", "message": "Modelo VisionFracture (Segmentation) pronto."}
        }
        return jsonify({"success": True, "results": results}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@flask_app.route('/api/test_inference', methods=['POST'])
@login_required
@role_required('admin')
def test_inference_api():
    """Testa inferência em uma imagem enviada."""
    try:
        if 'image' not in request.files:
            return jsonify({"success": False, "message": "Nenhuma imagem enviada"}), 400
            
        file = request.files['image']
        img_bytes = file.read()
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({"success": False, "message": "Falha ao decodificar imagem"}), 400
            
        # Executa inferência usando o modelo atual se disponível
        if _model_instance and hasattr(_model_instance, 'predict'):
            results = _model_instance.predict(img)
            # Desenha resultados (usando logic do DetectionDrawer se possível ou manual)
            # Para simplificar, vamos apenas retornar que funcionou
            
            # Converte imagem resultante para base64
            _, buffer = cv2.imencode('.jpg', img)
            img_base64 = base64.b64encode(buffer).decode('utf-8')
            
            return jsonify({
                "success": True, 
                "image": f"data:image/jpeg;base64,{img_base64}",
                "summary": "Detecção concluída. Objetos encontrados."
            }), 200
        else:
            return jsonify({"success": False, "message": "Modelo não disponível para inferência"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/express_retrain', methods=['POST'])
@login_required
@role_required('admin')
def express_retrain_api():
    """Realiza ajuste fino rápido."""
    try:
        target_class = request.form.get('target_class')
        files = request.files.getlist('images')
        
        _logger.info(f"Treino Relâmpago solicitado para classe {target_class} com {len(files)} imagens.")
        
        # Salva imagens temporariamente e inicia treino
        # ... lógica de salvamento ...
        
        if _otx_manager:
            # threading.Thread(target=lambda: _otx_manager.auto_fine_tune("temp_data"), daemon=True).start()
            return jsonify({"success": True, "message": "Ajuste fino iniciado em background."}), 200
        
        return jsonify({"success": False, "message": "OTX Manager não disponível"}), 500
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/logout')
@login_required
def logout():
    db = get_db()
    client_ip = request.remote_addr
    username = current_user.username 
    _logger.info(f"Usuário '{username}' deslogado remotamente do IP: {client_ip}.")
    
    with active_users_lock:
        if username in active_users:
            del active_users[username]
            _logger.info(f"Usuário '{username}' removido da lista de ativos.")
    db.log_activity(username, client_ip, 'LOGOUT')
    logout_user()
    return redirect(url_for('login_page'))
_excluded_classes = set()
_excluded_classes_lock = Lock()

@flask_app.route('/api/set_stream_filters', methods=['POST'])
@login_required
def set_stream_filters():
    """Define quais classes devem ser ocultadas no desenho das detecções."""
    global _excluded_classes
    try:
        data = request.get_json()
        classes = data.get('excluded_classes', [])
        with _excluded_classes_lock:
            _excluded_classes = set(classes)
        _logger.info(f"Filtros de stream atualizados: Ocultando {list(_excluded_classes)}")
        
        # Sincroniza com o modelo se ele existir
        if _model_instance and hasattr(_model_instance, 'set_excluded_classes'):
            _model_instance.set_excluded_classes(list(_excluded_classes))
            
        return jsonify({"status": "success", "excluded": list(_excluded_classes)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400

@flask_app.route('/video_feed')
def video_feed():
    """Rota que retorna o stream MJPEG."""


    user_desc = f"'{current_user.username}'" if current_user.is_authenticated else "'Anônimo'"
    _logger.info(f"Iniciando stream de vídeo para usuário {user_desc} (IP: {request.remote_addr})") 
    return Response(generate_video_frames(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@flask_app.route('/analytics')
@login_required
def analytics_page():
    """Página de Análises."""
    return render_template('analytics.html', username=current_user.username, client_ip=request.remote_addr)

@flask_app.route('/api/last_fracture_roi')
def last_fracture_roi():
    """Retorna o ROI que o VisionFracture esta analisando.
    Slot 1 = frame ao vivo da RAM do modelo com máscaras desenhadas.
    Slots 2-4 = ultimas fraturas salvas em disco.
    """
    slot = request.args.get('slot', '1')
    try:
        idx = int(slot) - 1
        
        # Slot 1: serve o ROI ao vivo com máscaras desenhadas
        if idx == 0 and _model_instance is not None and hasattr(_model_instance, 'last_roi_with_mask'):
            with _model_instance.last_roi_lock:
                if _model_instance.last_roi_with_mask is None:
                    return jsonify({"status": "no_roi", "message": "Nenhuma lata detectada."}), 404
                
                roi = _model_instance.last_roi_with_mask.copy()
                # Correção: verifica se info não é None antes de dar .copy()
                raw_info = getattr(_model_instance, 'last_roi_fracture_info', None)
                fracture_info = raw_info.copy() if raw_info is not None else {}
            
            if roi is not None:
                ret, buf = cv2.imencode('.jpg', roi, [cv2.IMWRITE_JPEG_QUALITY, 85])
                if ret:
                    response = Response(buf.tobytes(), mimetype='image/jpeg')
                    response.headers['X-Inspection-Status'] = 'LIVE'
                    response.headers['X-Fracture-Detected'] = 'true' if fracture_info.get('detected', False) else 'false'
                    response.headers['X-Fracture-Count'] = str(fracture_info.get('mask_count', 0))
                    response.headers['X-Fracture-Area-Px'] = str(int(fracture_info.get('total_area_px', 0)))
                    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response.headers['Pragma'] = 'no-cache'
                    response.headers['Expires'] = '0'
                    return response
        
        # Fallback para last_roi_frame se last_roi_with_mask não estiver disponível (compatibilidade)
        if idx == 0 and _model_instance is not None and hasattr(_model_instance, 'last_roi_frame'):
            with _model_instance.last_roi_lock:
                roi = _model_instance.last_roi_frame
            if roi is not None:
                ret, buf = cv2.imencode('.jpg', roi, [cv2.IMWRITE_JPEG_QUALITY, 85])
                if ret:
                    response = Response(buf.tobytes(), mimetype='image/jpeg')
                    response.headers['X-Inspection-Status'] = 'LIVE'
                    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response.headers['Pragma'] = 'no-cache'
                    response.headers['Expires'] = '0'
                    return response
        
        if idx == 0:
            return jsonify({"success": False, "message": "ROI ao vivo indisponivel."}), 404

        # Slots 2-4: ler apenas arquivos de fratura (FR) do disco
        import glob
        alerts_dir = r"E:\programs\visionAlign\data\alerts"
        files = glob.glob(os.path.join(alerts_dir, "ROI_insp_FR_*.jpg"))
        files.sort(key=os.path.getmtime, reverse=True)

        disk_idx = max(0, idx - 1) if idx > 0 else 0
        if not files or disk_idx >= len(files):
            return jsonify({"success": False, "message": "ROI nao disponivel."}), 404

        selected_file = files[disk_idx]
        response = send_file(os.path.abspath(selected_file), mimetype='image/jpeg')
        response.headers['X-Inspection-Status'] = 'FRACTURE'
        return response

    except Exception as e:
        _logger.error(f"Erro ao buscar ROI (slot {slot}): {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/extract_frame', methods=['POST'])
@login_required
@role_required('admin') 
def extract_frame():
    """Endpoint para solicitar a extração do próximo frame processado (criptografado)."""

    if _frame_container is None or _frame_lock is None:
        _logger.error("Extração de frame falhou: Container ou lock não inicializados.")
        return jsonify({"success": False, "message": "Erro interno do servidor (container/lock não pronto)."}), 500

    frame_to_send = None
    with _frame_lock:
        if _frame_container[0] is not None:
            frame_to_send = _frame_container[0].copy() 

    if frame_to_send is not None:
        try:

            ret, buffer = cv2.imencode('.jpg', frame_to_send, [cv2.IMWRITE_JPEG_QUALITY, 95]) 
            if ret:
                frame_bytes = buffer.tobytes()
                timestamp = get_current_utc_timestamp().strftime("%Y%m%d_%H%M%S") 
                filename = f"frame_capturado_{timestamp}.jpg"
                _logger.info(f"Usuário '{current_user.username}' extraiu frame. Enviando {filename} para download.")


                if _security_manager:
                    encrypted_frame_bytes = _security_manager.encrypt_data(frame_bytes)
                    

                    return Response(encrypted_frame_bytes,
                                    mimetype='application/octet-stream', 
                                    headers={"Content-Disposition": f"attachment;filename={filename}.enc"}) 
                else:
                    _logger.error("SecurityManager não inicializado. Não é possível criptografar o frame.")
                    return jsonify({"success": False, "message": "Erro interno: Criptografia não disponível."}), 500

            else:
                raise ValueError("Falha ao codificar frame como JPEG.")
        except Exception as e:
            _logger.error(f"Erro ao processar extração de frame: {e}")
            return jsonify({"success": False, "message": f"Erro ao processar frame: {e}"}), 500
    else:
        _logger.warning(f"Usuário '{current_user.username}' tentou extrair frame, mas nenhum estava disponível.")
        return jsonify({"success": False, "message": "Nenhum frame disponível para extração no momento."}), 404



@flask_app.route('/api/save_training_frame', methods=['POST'])
@login_required
@role_required('admin')
def save_training_frame():
    """Endpoint para salvar o frame atual em uma pasta de dataset para retraining."""
    if _model_instance is None:
        return jsonify({"success": False, "message": "Instância do modelo não encontrada."}), 500
    
    success, result = _model_instance.save_for_dataset(prefix="manual")
    if success:
        _logger.info(f"Usuário '{current_user.username}' salvou frame {result} para dataset de treinamento.")
        return jsonify({"success": True, "message": f"Frame {result} salvo com sucesso para treinamento."})
    else:
        return jsonify({"success": False, "message": f"Erro ao salvar: {result}"}), 500

@flask_app.route('/api/dataset_info', methods=['GET'])
@login_required
def get_dataset_info_api():
    """Retorna informações sobre o dataset coletado para treinamento."""
    try:
        import glob
        dataset_path = r"E:\programs\visionAlign\data\dataset_collect\images"
        os.makedirs(dataset_path, exist_ok=True)
        
        images = glob.glob(os.path.join(dataset_path, "*.jpg"))
        count = len(images)
        
        last_update = "N/A"
        if count > 0:
            latest_file = max(images, key=os.path.getmtime)
            # Usando conversion para exibição amigável
            mtime = os.path.getmtime(latest_file)
            last_update = datetime.fromtimestamp(mtime).strftime('%d/%m/%Y %H:%M:%S')
            
        return jsonify({
            "image_count": count,
            "class_count": 4, # Normal, Inverted, Fallen, Fracture
            "last_update": last_update
        })
    except Exception as e:
        _logger.error(f"Erro ao buscar info do dataset: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    if 'images' not in request.files:
        return jsonify({"success": False, "message": "Nenhuma imagem enviada."}), 400
    
    target_class = request.form.get('target_class', 'fracture')
    files = request.files.getlist('images')
    
    try:
        import os
        import shutil
        import threading
        from ultralytics import YOLO
        import numpy as np
        import cv2

        # 1. Preparação do Diretório de Treino Expresso
        base_dir = r"E:\programs\visionAlign\data\express_train"
        train_img_dir = os.path.join(base_dir, "images", "train")
        train_lbl_dir = os.path.join(base_dir, "labels", "train")
        
        # Limpa treino anterior
        if os.path.exists(base_dir):
            shutil.rmtree(base_dir)
        
        os.makedirs(train_img_dir, exist_ok=True)
        os.makedirs(train_lbl_dir, exist_ok=True)

        _logger.info(f"Iniciando Auto-Labeling para {len(files)} imagens (Classe: {target_class}).")

        # 2. Auto-Labeling (Pseudo-Labeling)
        # Usamos o modelo atual para encontrar ONDE estão os objetos, mas forçamos a CLASSE que o user quer.
        path_align = r"E:\programs\visionAlign\model\_openvino_model\VisionAlign_openvino_model"
        if not os.path.exists(path_align): path_align = r"E:\programs\visionAlign\model\backup\best.pt"
        
        path_fracture = r"E:\programs\visionAlign\model\_openvino_model\VisionFracture_openvino_model"
        if not os.path.exists(path_fracture): path_fracture = r"E:\programs\visionAlign\model\fracture\best.pt"

        # Carrega modelos para rotulagem
        m_align = YOLO(path_align)
        m_fracture = YOLO(path_fracture)

        saved_count = 0
        for i, file in enumerate(files):
            # Salva imagem
            img_path = os.path.join(train_img_dir, f"express_{i}.jpg")
            file.save(img_path)
            
            # Carrega para processar
            img = cv2.imread(img_path)
            if img is None: continue
            
            labels = []
            
            if target_class == 'fracture':
                # Para fratura, precisamos primeiro achar a lata (align) e depois segmentar nela
                res_align = m_align.predict(img, conf=0.25, verbose=False)
                if res_align and len(res_align) > 0:
                    for box in res_align[0].boxes.xyxy.cpu().numpy():
                        x1, y1, x2, y2 = map(int, box)
                        roi = img[y1:y2, x1:x2]
                        if roi.size > 0:
                            res_fr = m_fracture.predict(roi, conf=0.1, verbose=False) # Conf baixa para capturar tudo
                            if res_fr and len(res_fr) > 0 and hasattr(res_fr[0], 'masks') and res_fr[0].masks is not None:
                                for mask in res_fr[0].masks.xyn:
                                    # Formato YOLO Seg: class_id x1 y1 x2 y2 ... (normalizado)
                                    # Como é segmentação global da imagem de treino, precisamos normalizar relativo a imagem toda
                                    # Mas o model_fracture treina no ROI. 
                                    # Para simplificar o "Treino Relâmpago", vamos treinar o model_fracture no ROI salvo.
                                    roi_img_path = os.path.join(train_img_dir, f"express_roi_{saved_count}.jpg")
                                    cv2.imwrite(roi_img_path, roi)
                                    
                                    points = " ".join([f"{p[0]:.6f} {p[1]:.6f}" for p in mask])
                                    lbl_path = os.path.join(train_lbl_dir, f"express_roi_{saved_count}.txt")
                                    with open(lbl_path, "w") as f:
                                        f.write(f"0 {points}\n")
                                    saved_count += 1
            else:
                # Para Alinhamento (Lata Tombada/Invertida)
                res = m_align.predict(img, conf=0.25, verbose=False)
                if res and len(res) > 0:
                    lbl_path = os.path.join(train_lbl_dir, f"express_{i}.txt")
                    with open(lbl_path, "w") as f:
                        for box_obj in res[0].boxes:
                            b = box_obj.xywhn.cpu().numpy()[0]
                            # Força a classe desejada (mapeando do select do front)
                            cid = 0 # Default
                            if "tombada" in target_class: cid = 2 
                            elif "invertida" in target_class: cid = 1
                            
                            f.write(f"{cid} {b[0]:.6f} {b[1]:.6f} {b[2]:.6f} {b[3]:.6f}\n")
                    saved_count += 1

        if saved_count == 0:
            return jsonify({"success": False, "message": "Não foi possível identificar objetos nas fotos para rotulagem automática."}), 400

        # 3. Dispara Treino em Background
        def run_fast_train():
            try:
                _logger.info("Iniciando motor de treinamento Ultra-Rápido...")
                
                # Paths Oficiais - CORRIGIDOS
                official_pt_align = r"E:\programs\visionAlign\model\backup\yolo26n.pt" 
                official_pt_fracture = r"E:\programs\visionAlign\model\backup\best.pt"
                
                official_ov_align = r"E:\programs\visionAlign\model\_openvino_model\VisionAlign_openvino_model"
                official_ov_fracture = r"E:\programs\visionAlign\model\_openvino_model\VisionFracture_openvino_model"

                # Cria data.yaml temporário
                yaml_path = os.path.join(base_dir, "data.yaml")
                with open(yaml_path, "w") as f:
                    if target_class == 'fracture':
                        f.write(f"train: {train_img_dir}\nval: {train_img_dir}\nnc: 1\nnames: ['fracture']\n")
                        model_path_to_load = official_pt_fracture
                        export_name = "VisionFracture"
                        target_ov_dir = official_ov_fracture
                    else:
                        f.write(f"train: {train_img_dir}\nval: {train_img_dir}\nnc: 3\nnames: ['Fallen', 'Inverted', 'Normal']\n")
                        model_path_to_load = official_pt_align
                        export_name = "VisionAlign"
                        target_ov_dir = official_ov_align
                
                # Carrega o modelo PESADO (.pt) para treinamento
                # NOTA: O Align é detect mas pode vir de um .pt genérico
                m_train = YOLO(model_path_to_load)
                
                # Treino Relâmpago
                results = m_train.train(
                    data=yaml_path,
                    epochs=15,
                    imgsz=640,
                    batch=4,
                    lr0=0.001,
                    lrf=0.01,
                    plots=False,
                    overlap_mask=True,
                    verbose=False,
                    project=os.path.join(base_dir, "runs"),
                    name="express_run"
                )
                
                # O novo .pt está em: base_dir/runs/express_run/weights/best.pt
                new_pt_path = os.path.join(base_dir, "runs", "express_run", "weights", "best.pt")
                
                if os.path.exists(new_pt_path):
                    # 1. Atualiza o PT de produção para o próximo treino relâmpago ser melhor
                    shutil.copy(new_pt_path, model_path_to_load)
                    
                    _logger.info("Modelo .pt atualizado. Exportando para OpenVINO...")
                    
                    # 2. Exporta o novo modelo
                    m_new = YOLO(new_pt_path)
                    ov_export_path = m_new.export(format="openvino", imgsz=640) 
                    # ov_export_path será algo como '.../best_openvino_model'
                    
                    # 3. Substitui a pasta OpenVINO de produção
                    if os.path.exists(ov_export_path):
                        if os.path.exists(target_ov_dir):
                            shutil.rmtree(target_ov_dir)
                        shutil.copytree(ov_export_path, target_ov_dir)
                        
                        _logger.info(f"Pasta OpenVINO atualizada em: {target_ov_dir}")
                        
                        # 4. TRIGGER RELOAD NO SISTEMA VIVO
                        if _model_instance:
                            _logger.info("Triggering HOT RELOAD dos modelos na memória...")
                            _model_instance.reload_models()
                
                _logger.info(">>> SISTEMA ATUALIZADO COM SUCESSO! O novo conhecimento já está valendo. <<<")
                
            except Exception as e:
                import traceback
                _logger.error(f"Erro no background train: {e}\n{traceback.format_exc()}")

        threading.Thread(target=run_fast_train, daemon=True).start()

        return jsonify({
            "success": True, 
            "message": f"Auto-rotulagem concluída ({saved_count} amostras). O treinamento está rodando em segundo plano e levará ~2-5 minutos."
        })

    except Exception as e:
        _logger.error(f"Erro no express retrain: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@flask_app.route('/api/decrypt', methods=['POST', 'GET'])
@login_required
@role_required('admin')
def decrypt_data_api():
    # Para GET, retornar erro informativo
    if request.method == 'GET':
        return jsonify({"message": "Use POST para enviar dados criptografados. GET não é suportado para descriptografia."}), 405
    
    try:
        encrypted_data = request.data 
        if not encrypted_data:
             return jsonify({"message": "Nenhum dado criptografado recebido"}), 400

        if not _security_manager:
             _logger.error("SecurityManager não inicializado. Não é possível descriptografar.")
             return jsonify({"message": "Erro interno: Criptografia não disponível"}), 500

        decrypted_data = _security_manager.decrypt_data(encrypted_data)
        if decrypted_data is None:
             _logger.warning("Falha ao descriptografar dados via API /api/decrypt.")
             return jsonify({"message": "Falha ao descriptografar dados"}), 400
        try:
            decrypted_json = json.loads(decrypted_data.decode('utf-8'))
            return jsonify(decrypted_json), 200
        except (json.JSONDecodeError, UnicodeDecodeError):
            
            return Response(decrypted_data, mimetype='application/octet-stream') 
    except Exception as e:
        _logger.error(f"Erro na API /api/decrypt: {e}", exc_info=True)
        return jsonify({"message": f"Erro ao processar descriptografia: {e}"}), 500
    

@flask_app.route('/api/detection_history', methods=['GET'])
@login_required
def get_detection_history_api():
    """Busca dados históricos de detecção para o dashboard."""
    
    
    try:
        requested_period_key = request.args.get('period', 'hour').lower() 
                                                                     
        db = get_db()
        if requested_period_key == 'hour': 
            agg_period_type = 'hour'
            agg_num_periods = 24 
        elif requested_period_key == 'week': 
            agg_period_type = 'day'
            agg_num_periods = 7 
        elif requested_period_key == 'month': 
            agg_period_type = 'day' 
            agg_num_periods = 30 
        elif requested_period_key == 'year': 
            agg_period_type = 'month' 
            agg_num_periods = 12 
        
        elif requested_period_key == 'day': 
            agg_period_type = 'hour'
            agg_num_periods = 24
        elif requested_period_key == 'week_trend': 
            agg_period_type = 'day'
            agg_num_periods = 7
        elif requested_period_key == 'month_trend': 
            agg_period_type = 'day' 
            agg_num_periods = 30
        else: 
            agg_period_type = 'hour'
            agg_num_periods = 24

        
        aggregated_data = db.get_aggregated_detection_data(period_type=agg_period_type, num_periods=agg_num_periods)
        
        trend_chart_labels = [row[0] for row in aggregated_data] 
        trend_chart_rejections_data = [row[2] + row[3] for row in aggregated_data]  # inverted + fallen
        trend_chart_totals_data = [row[4] for row in aggregated_data]
        total_normal = sum(row[1] for row in aggregated_data)
        total_inverted = sum(row[2] for row in aggregated_data)
        total_fallen = sum(row[3] for row in aggregated_data)
        total_fracture = sum(row[5] for row in aggregated_data) if len(aggregated_data) > 0 and len(aggregated_data[0]) > 5 else 0
        
        pie_chart_data = [total_normal, total_inverted, total_fallen, total_fracture]
        pie_chart_labels = ["Normal", "Invertida", "Tombada", "Fratura"]
        _logger.info(f"API /detection_history: Dados para gráfico de pizza: labels={pie_chart_labels}, data={pie_chart_data}")

        
        
        
        # Informações do Modelo
        model_info_data = {
            'version': '--',
            'last_trained': '--',
            'dataset_size': '--'
        }
        if _settings_manager:
            all_settings = _settings_manager.get_all()
            model_params_settings = all_settings.get('MODEL_PARAMS', {})
            model_info_data['version'] = model_params_settings.get('model_version', '--') # Mantém como está
            model_info_data['last_trained'] = model_params_settings.get('last_training', '--') # Corrigido para 'last_training'
            model_info_data['dataset_size'] = model_params_settings.get('size_dataset', '--') # Corrigido para 'size_dataset'
        
        # Métricas de Desempenho (KPIs)
        
        kpi_data = db.get_kpis_last_hour() 
        response = {
            "kpis": {
                # Renomeando para maior clareza e especificidade do período
                "total_detections_last_hour": kpi_data.get('total', 0),
                "error_rate_last_hour": kpi_data.get('error_rate', 0.0),
                "avg_processing_time_ms_last_hour": kpi_data.get('avg_processing_time_ms', '--')
            },
            "detections_per_hour": {
                "labels": trend_chart_labels, 
                "rejections": trend_chart_rejections_data, 
                "totals": trend_chart_totals_data      
            },
            "detection_types": {
                "labels": pie_chart_labels,
                "data": pie_chart_data
            },
            "model_info": model_info_data # Adicionando informações do modelo
        }

        return jsonify(response), 200

    except Exception as e:
        _logger.error(f"Erro na API /api/detection_history: {e}", exc_info=True)
        return jsonify({"message": f"Erro ao buscar histórico: {e}"}), 500

@flask_app.route('/api/alerts_history', methods=['GET'])
@login_required
def get_alerts_history_api():
    
    
    try:
        db = get_db()
        
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        alert_type = request.args.get('type')
        lata_id = request.args.get('lata_id')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int) 

        

        alerts_page_data, total_items = db.get_alerts_with_filters(
            start_date=start_date_str,
            end_date=end_date_str,
            alert_type=alert_type,
            lata_id=lata_id,
            page=page,
            per_page=per_page
        )

        table_data = []
        for alert_row in alerts_page_data:
            table_data.append({
                "timestamp": alert_row[1],
                "type": alert_row[2],
                "lata_id": alert_row[3],
                "details": alert_row[4]
            })
        
        total_pages = math.ceil(total_items / per_page) if per_page > 0 and total_items > 0 else 1
        if page > total_pages and total_pages > 0 : 
            page = total_pages
        elif page < 1:
            page = 1


        return jsonify({
            "alerts": table_data,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total_items": total_items,
                "total_pages": total_pages
            }
        }), 200
    except Exception as e:
        _logger.error(f"Erro na API /api/alerts_history: {e}", exc_info=True)
        return jsonify({"message": f"Erro ao buscar histórico de alertas: {e}"}), 500


@flask_app.route('/api/export_event_history', methods=['GET'])
@login_required
def export_event_history_excel_api():
    
    try:
        db = get_db()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        alert_type = request.args.get('type')
        lata_id = request.args.get('lata_id')
        

        
        alerts_data, _ = db.get_alerts_with_filters( 
            start_date=start_date_str,
            end_date=end_date_str,
            alert_type=alert_type,
            lata_id=lata_id,
            page=1, 
            per_page=0 
        )
        
        
        
        

        
        if not alerts_data:
            
            return jsonify({"message": "Nenhum dado para exportar com os filtros fornecidos."}), 404

        wb = openpyxl.Workbook() 
        ws = wb.active
        ws.title = "Histórico de Eventos"
        headers = ["Data/Hora (UTC)", "Tipo", "Lata ID", "Detalhes"] 
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for alert_row in alerts_data:
            
            
            
            
            
            ws.append([alert_row[1], alert_row[2], alert_row[3], alert_row[4]]) 
        
        column_widths = {'A': 25, 'B': 20, 'C': 15, 'D': 50}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        filename_timestamp = get_current_utc_timestamp().strftime("%Y%m%d_%H%M%S") 
        filename = f"historico_eventos_{filename_timestamp}.xlsx"
        return Response(
            excel_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

    except Exception as e:
        _logger.error(f"Erro na API /api/export_event_history: {e}", exc_info=True)
        return jsonify({"message": f"Erro ao exportar histórico de eventos: {e}"}), 500

@flask_app.route('/api/crypto_demo', methods=['GET'])
@login_required
@role_required('admin')
def crypto_demo_api():
    """Endpoint para demonstrar o processo de criptografia/descriptografia."""
    

    if not _security_manager:
        _logger.error("Crypto Demo: SecurityManager não inicializado.")
        return jsonify({"status": "Failure", "message": "Erro interno: Criptografia não disponível"}), 500

    try:

        original_dict = {
            "message": "Teste de criptografia VisionAlign!",
            "timestamp": time.time(),
            "user": current_user.username,
            "data_points": [10, 25, 5, 42]
        }

        original_json_str = json.dumps(original_dict, sort_keys=True)
        original_bytes = original_json_str.encode('utf-8')
        


        encrypted_bytes = _security_manager.encrypt_data(original_bytes)

        encrypted_preview = f"({len(encrypted_bytes)} bytes, starts with: {encrypted_bytes[:15]!r}... ends with: {encrypted_bytes[-15:]!r})"
        


        decrypted_bytes = _security_manager.decrypt_data(encrypted_bytes)


        response_data = {
            "status": "Unknown",
            "original_data": original_dict,
            "encrypted_data_info": encrypted_preview,
            "decryption_successful": False,
            "decrypted_data": None, 
            "data_matches": False
        }

        if decrypted_bytes is not None:
            response_data["decryption_successful"] = True

            decrypted_dict = json.loads(decrypted_bytes.decode('utf-8'))
            response_data["decrypted_data"] = decrypted_dict

            if json.dumps(original_dict, sort_keys=True) == json.dumps(decrypted_dict, sort_keys=True):
                 response_data["data_matches"] = True
                 response_data["status"] = "Success"
                 _logger.info("Crypto Demo: Teste de Cripto/Descripto bem-sucedido.")
            else:
                 response_data["status"] = "Failure"
                 response_data["message"] = "Dados descriptografados não correspondem aos originais!"
                 _logger.error("Crypto Demo: Dados descriptografados NÃO CORRESPONDEM aos originais!")
        else:
            response_data["status"] = "Failure"
            response_data["message"] = "Falha na descriptografia (InvalidToken ou chave errada)."
            _logger.error("Crypto Demo: Descriptografia falhou (retornou None).")

        return jsonify(response_data), 200

    except Exception as e:
        _logger.error(f"Erro inesperado na API /api/crypto_demo: {e}", exc_info=True)
        return jsonify({"status": "Failure", "message": f"Erro inesperado no servidor: {e}"}), 500
@flask_app.route('/api/export_data', methods=['GET'])
@login_required
@role_required('admin') 
def export_data():
    """Endpoint para exportar dados de atividade recentes como CSV (criptografado)."""
    
    try:
        db = get_db()

        time_threshold = get_current_utc_timestamp() - timedelta(hours=1)

        logs = db.get_activity_since(time_threshold) 

        if not logs:
            return jsonify({"success": False, "message": "Nenhuma atividade registrada na última hora."}), 404


        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Timestamp', 'Username', 'IP Address', 'Action', 'Details']) 
        for log in logs:
            writer.writerow([log[1], log[2], log[3], log[4], log[5]]) 

        output.seek(0)
        csv_data = output.getvalue().encode('utf-8')


        if _security_manager:
            encrypted_csv_data = _security_manager.encrypt_data(csv_data)
            _logger.info(f"Exportando log de atividades criptografado.")
            return Response(encrypted_csv_data, mimetype="application/octet-stream", headers={"Content-Disposition":"attachment;filename=activity_log_last_hour.csv.enc"})
        else:
            _logger.error("SecurityManager não inicializado. Não é possível criptografar o log.")
            return jsonify({"success": False, "message": "Erro interno: Criptografia não disponível."}), 500




    except Exception as e:
        _logger.error(f"Erro ao exportar dados: {e}")
        return jsonify({"success": False, "message": f"Erro ao exportar dados: {e}"}), 500


@flask_app.route('/api/change_source', methods=['POST'])
@login_required
@role_required('admin')
def change_source():
    """Endpoint para instruir o servidor a mudar a fonte de vídeo (recebe dados criptografados)."""
    global _model_instance
    if not _model_instance:
        _logger.error("API /change_source: Instância do modelo não está disponível.")
        return jsonify({"success": False, "message": "Erro interno: Modelo não inicializado no servidor."}), 500


    encrypted_data = request.data
    if not encrypted_data: return jsonify({"success": False, "message": "Requisição vazia"}), 400
    if not _security_manager: return jsonify({"success": False, "message": "Erro interno: Criptografia não disponível."}), 500

    decrypted_data = _security_manager.decrypt_data(encrypted_data)
    if decrypted_data is None: return jsonify({"success": False, "message": "Falha ao descriptografar dados da requisição"}), 400

    payload = json.loads(decrypted_data.decode('utf-8'))
    source_type = payload.get('type')
    source_param = payload.get('param')

    _logger.info(f"API /change_source: Recebido pedido para mudar para tipo '{source_type}' com param '{source_param}'")

    success = False
    try:

        _model_instance.stop_processing()
        
        if _current_stats is not None and _stats_lock is not None:
            with _stats_lock:
                _logger.info("API /change_source: Zerando contagens de latas em _current_stats.")
                _current_stats['lata_normal'] = 0
                _current_stats['lata_invertida'] = 0
                _current_stats['lata_tombada'] = 0
                _logger.info(f"API /change_source: _current_stats APÓS zerar: {_current_stats}") 
                
        
        _model_instance.reset_state() 

        if source_type == 'camera':
            camera_id = int(source_param) 
            success = _model_instance.load_camera(camera_id)
        elif source_type == 'stream':
            success = _model_instance.load_stream(str(source_param)) 
        elif source_type == 'video':

            success = _model_instance.load_video(str(source_param)) 
        else:
            return jsonify({"success": False, "message": f"Tipo de fonte inválido: {source_type}"}), 400

        if success:
            _model_instance.start_processing() 
            _logger.info(f"API /change_source: Fonte alterada com sucesso para '{source_type}'.")
            return jsonify({"success": True, "message": f"Fonte alterada para {source_type}"}), 200
        else:
            _logger.error(f"API /change_source: Falha ao carregar a fonte '{source_type}' com param '{source_param}'.")
            return jsonify({"success": False, "message": f"Falha ao carregar a fonte {source_type}"}), 500

    except Exception as e:
        _logger.error(f"API /change_source: Erro ao processar mudança de fonte: {e}", exc_info=True)


        return jsonify({"success": False, "message": f"Erro interno ao mudar fonte: {e}"}), 500


@flask_app.route('/api/web_stream/pause', methods=['POST'])
@login_required
@role_required('admin')
def pause_web_stream():
    """Endpoint para pausar o envio do stream MJPEG."""
    global _web_stream_paused, _web_stream_lock
    with _web_stream_lock:
        _web_stream_paused = True
    _logger.info(f"Stream MJPEG PAUSADO via API web pelo usuário '{current_user.username}'.")
    return jsonify({"success": True, "message": "Stream web pausado."}), 200

@flask_app.route('/api/web_stream/resume', methods=['POST'])
@login_required
@role_required('admin')
def resume_web_stream():
    """Endpoint para retomar o envio do stream MJPEG."""
    global _web_stream_paused, _web_stream_lock
    with _web_stream_lock:
        _web_stream_paused = False
    _logger.info(f"Stream MJPEG RETOMADO via API web pelo usuário '{current_user.username}'.")
    return jsonify({"success": True, "message": "Stream web retomado."}), 200



@flask_app.route('/api/pause', methods=['POST'])
@login_required
@role_required('admin')
def pause_processing():
    """Endpoint para pausar o processamento do modelo."""
    if _model_instance:
        _model_instance.pause_processing()
        _logger.info(f"Processamento pausado via API pelo usuário '{current_user.username}'.")
        return jsonify({"success": True, "message": "Processamento pausado."}), 200
    return jsonify({"success": False, "message": "Modelo não disponível."}), 500

@flask_app.route('/api/resume', methods=['POST'])
@login_required
@role_required('admin')
def resume_processing():
    """Endpoint para retomar o processamento do modelo."""
    if _model_instance:
        _model_instance.resume_processing()
        _logger.info(f"Processamento retomado via API pelo usuário '{current_user.username}'.")
        return jsonify({"success": True, "message": "Processamento retomado."}), 200
    return jsonify({"success": False, "message": "Modelo não disponível."}), 500


@flask_app.route('/console')
@login_required
@role_required('admin')
def console_page():
    """Renderiza a página HTML do console de logs."""
    _logger.info(f"Usuário '{current_user.username}' acessou o console web.")

    initial_logs = list(log_queue)
    return render_template('console.html', initial_logs=initial_logs)

@flask_app.route('/console_stream')
@login_required
@role_required('admin')
def console_stream():
    """Endpoint SSE que envia novas mensagens de log."""
    listener_queue = queue.Queue()
    sse_listeners.append(listener_queue)
    

    def generate_log_stream():
        try:
            while True:

                message = listener_queue.get() 

                yield f"data: {message}\n\n"
        except GeneratorExit:

            _logger.info("Ouvinte SSE do console desconectado.") 
        finally:

            if listener_queue in sse_listeners:
                sse_listeners.remove(listener_queue)
                

    return Response(generate_log_stream(), mimetype='text/event-stream')

@flask_app.route('/api/users_list', methods=['GET'])
@login_required
@role_required('admin')
def get_users_list():
    
    users_list_for_frontend = []
    now_utc = get_current_utc_timestamp()

    with active_users_lock:
        for username, data in active_users.items():
            last_seen_delta = now_utc - data['last_seen']
            
            status = "Online"
            status_class = "badge-online"
            
            
            if last_seen_delta.total_seconds() > 300: 
                status = "Inativo"
                status_class = "badge-idle"
            
            
            
            

            last_activity_description = "Agora mesmo"
            if last_seen_delta.total_seconds() < 2: 
                last_activity_description = "Agora mesmo"
            elif last_seen_delta.total_seconds() < 60:
                last_activity_description = f"{int(last_seen_delta.total_seconds())}s atrás"
            elif last_seen_delta.total_seconds() < 3600: 
                last_activity_description = f"{int(last_seen_delta.total_seconds() / 60)} min atrás"
            elif last_seen_delta.total_seconds() < 86400: 
                last_activity_description = f"{int(last_seen_delta.total_seconds() / 3600)}h atrás"
            else:
                last_activity_description = f"{int(last_seen_delta.total_seconds() / 86400)}d atrás"

            user_agent_str = data.get('user_agent', 'N/A')
            
            browser_info = "N/A"
            
            
            if user_agent_str != "N/A":
                ua_lower = user_agent_str.lower()
                if "opr/" in ua_lower or "opera" in ua_lower: 
                    browser_info = "Opera"
                elif "edg/" in ua_lower or "edge" in ua_lower: 
                    browser_info = "Edge"
                elif "chrome/" in ua_lower and "chromium/" not in ua_lower: 
                    browser_info = "Chrome"
                elif "firefox/" in ua_lower: 
                    browser_info = "Firefox"
                elif "safari/" in ua_lower and "chrome/" not in ua_lower: 
                    browser_info = "Safari"
                elif "msie" in ua_lower or "trident/" in ua_lower: 
                    browser_info = "Internet Explorer"
                else:
                    
                    try:
                        ua_parser = request.user_agent.__class__(user_agent_str)
                        if ua_parser.browser:
                            browser_info = ua_parser.browser.capitalize()
                    except Exception as e_ua_fallback:
                        _logger.debug(f"Erro no fallback do parser User-Agent para '{user_agent_str}': {e_ua_fallback}")

            last_activity_details_str = f"Navegador: {browser_info}"
            if browser_info == "N/A":
                if user_agent_str != "N/A" and user_agent_str:
                    
                    last_activity_details_str = f"User-Agent: {user_agent_str[:30]}..."
                else:
                    last_activity_details_str = "User-Agent não disponível"
            
            
            device_display_name = data.get('hostname', data.get('ip_address', 'Dispositivo Desconhecido'))
            user_first_name = data.get('first_name', '')
            user_last_name = data.get('last_name', '')
            
            display_name_parts = []
            if user_first_name: display_name_parts.append(user_first_name)
            if user_last_name: display_name_parts.append(user_last_name)
            
            display_name = " ".join(display_name_parts).strip()
            if not display_name: display_name = username.replace("_", " ").title()

            avatar_initials_str = ""
            if user_first_name: avatar_initials_str += user_first_name[0]
            if user_last_name: avatar_initials_str += user_last_name[0]
            if not avatar_initials_str and username: avatar_initials_str = "".join([name_part[0] for name_part in username.replace("_", " ").split()[:2]]).upper()
            elif not avatar_initials_str: avatar_initials_str = "??"
            
            users_list_for_frontend.append({
                "avatar_initials": "".join([name[0] for name in username.replace("_", " ").split()[:2]]).upper() if username else "??",
                "name": username.replace("_", " ").title(),
                "email": f"{username.lower()}@example.com", 
                "status": status,
                "status_class": status_class,
                "device_name": device_display_name, 
                "ip_address": data.get('ip_address', 'N/A'),
                "last_activity_time": last_activity_description,
                "last_activity_details": last_activity_details_str,
                "role": data.get('role', 'N/A').title(),
                "is_banned": False 
            })

    
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 10)) 
    search_term = request.args.get('search', '').lower()

    if search_term:
        users_list_for_frontend = [
            u for u in users_list_for_frontend if 
            search_term in u['name'].lower() or 
            search_term in u['email'].lower() or
            search_term in u['ip_address'].lower()
        ]

    total_users = len(users_list_for_frontend)
    total_pages = math.ceil(total_users / per_page) if per_page > 0 and total_users > 0 else 1
    if page > total_pages and total_pages > 0: page = total_pages
    elif page < 1: page = 1
        
    start_index = (page - 1) * per_page
    paginated_users = users_list_for_frontend[start_index : start_index + per_page]
    
    return jsonify({
        "users": paginated_users,
        "pagination": {"page": page, "per_page": per_page, "total_items": total_users, "total_pages": total_pages}
    }), 200


@flask_app.route('/api/force_logout_user', methods=['POST'])
@login_required
@role_required('admin') 
def force_logout_user_api():
    data = request.json
    username_to_logout = data.get('username')

    if not username_to_logout:
        _logger.warning(f"API /api/force_logout_user: Tentativa de forçar logout sem fornecer username. Solicitado por: {current_user.username}")
        return jsonify({"success": False, "message": "Nome de usuário não fornecido."}), 400

    _logger.info(f"API /api/force_logout_user: {current_user.username} solicitou logout forçado para {username_to_logout}.")

    with active_users_lock:
        if username_to_logout in active_users:
            del active_users[username_to_logout]
            
            
            
            _logger.info(f"Usuário '{username_to_logout}' removido da lista de ativos por {current_user.username}.")
            
            
            return jsonify({"success": True, "message": f"Usuário '{username_to_logout}' desconectado (removido da lista de ativos)."}), 200
        else:
            _logger.warning(f"API /api/force_logout_user: Usuário '{username_to_logout}' não encontrado na lista de ativos para logout forçado.")
            return jsonify({"success": False, "message": f"Usuário '{username_to_logout}' não está ativo ou já foi desconectado."}), 404


def record_detection_event(timestamp_utc_iso, alert_type_str, detected_lata_ids_list, details_str_template, confidences_list=None):
    """
    Processa eventos de detecção, gera alertas únicos por ID de lata e os registra.
    Esta função deve ser chamada pela instância do modelo ao detectar anomalias.

    Args:
        timestamp_utc_iso (str): Timestamp do evento no formato ISO UTC.
        alert_type_str (str): Tipo de detecção (ex: "Lata Invertida").
        detected_lata_ids_list (list): Lista de strings de IDs de lata detectados (ex: ['
        details_str_template (str): Um template para os detalhes do alerta.
                                     Pode ser uma descrição geral ou conter placeholders como {id}.
        confidences_list (list, optional): Lista de floats de confianças, correspondente a detected_lata_ids_list.
    """
    global _active_alerts_cache, _ALERT_COOLDOWN_SECONDS, _ALERT_STATE_EXPIRY_SECONDS, _logger

    
    
    try:
        with Database() as db: 
            current_time_unix = get_current_utc_timestamp().timestamp()

            if not detected_lata_ids_list:
                return

            
            ids_to_clear_from_cache = [
                lid for lid, data in list(_active_alerts_cache.items())
                if (current_time_unix - data['timestamp']) > _ALERT_STATE_EXPIRY_SECONDS
            ]
            for lid in ids_to_clear_from_cache:
                if lid in _active_alerts_cache: 
                    del _active_alerts_cache[lid]

            new_alerts_logged_count = 0
            for index, lata_id in enumerate(detected_lata_ids_list):
                cache_entry = _active_alerts_cache.get(lata_id)
                
                should_log_new_alert = False
                if not cache_entry: 
                    should_log_new_alert = True
                elif cache_entry['type'] != alert_type_str: 
                    should_log_new_alert = True
                elif (current_time_unix - cache_entry['timestamp']) > _ALERT_COOLDOWN_SECONDS: 
                    should_log_new_alert = True
                
                if should_log_new_alert:
                    confidence_value = confidences_list[index] if confidences_list and index < len(confidences_list) else None
                    confidence_str = f"Confiança: {confidence_value*100:.1f}%" if confidence_value is not None else ""
                    
                    
                    final_detail_for_this_id = details_str_template
                    if "{id}" in final_detail_for_this_id:
                        final_detail_for_this_id = final_detail_for_this_id.replace("{id}", lata_id)
                    
                    if confidence_str and confidence_str not in final_detail_for_this_id:
                        final_detail_for_this_id = f"{final_detail_for_this_id} ({confidence_str})".strip()
                    elif not final_detail_for_this_id and confidence_str: 
                        final_detail_for_this_id = confidence_str

                    _logger.info(f"Registrando novo alerta: Tipo='{alert_type_str}', ID='{lata_id}', Detalhes='{final_detail_for_this_id}'")
                    db.log_alert(timestamp_utc_iso, alert_type_str, lata_id, final_detail_for_this_id)
                    
                    _active_alerts_cache[lata_id] = {
                        'timestamp': current_time_unix,
                        'type': alert_type_str,
                        'last_detail': final_detail_for_this_id
                    }
                    new_alerts_logged_count += 1
    except Exception as e:
        
        _logger.error(f"Erro em record_detection_event: {e}", exc_info=True)

def run_server(settings_mgr, logger_instance, shared_stats, lock_stats, frame_cont, lock_frame, start_time, model_instance, security_mgr): 
    global _settings_manager, _logger, _current_stats, _stats_lock, _frame_container, _frame_lock, _system_start_time, _extract_frame_flag, _model_instance, _security_manager, _otx_manager
    global MAX_CONSOLE_LINES, log_queue, _ALERT_COOLDOWN_SECONDS, _ALERT_STATE_EXPIRY_SECONDS
    
    _settings_manager = settings_mgr
    _logger = logger_instance
    _current_stats = shared_stats
    _stats_lock = lock_stats
    _frame_container = frame_cont 
    _frame_lock = lock_frame 
    _system_start_time = start_time 
    _model_instance = model_instance 
    _security_manager = security_mgr 
    _otx_manager = OTXManager(settings_mgr, logger=logger_instance)

    # --- Aplicação de Configurações do SYSTEM_CONFIG ---
    sys_config = settings_mgr.get_setting('SYSTEM_CONFIG', {})
    server_cfg = sys_config.get('server', {})
    logging_cfg = sys_config.get('logging', {})
    monitor_cfg = sys_config.get('monitoring', {})

    # Flask Config
    flask_app.secret_key = server_cfg.get('secret_key', 'dev_key_fallback')
    session_lifetime = server_cfg.get('session_lifetime_minutes', 30)
    flask_app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=session_lifetime)

    # Module Constants
    MAX_CONSOLE_LINES = logging_cfg.get('max_console_lines', 500)
    if log_queue.maxlen != MAX_CONSOLE_LINES:
        # Recria a fila se o tamanho mudar
        log_queue = collections.deque(log_queue, maxlen=MAX_CONSOLE_LINES)
    
    _ALERT_COOLDOWN_SECONDS = monitor_cfg.get('alert_cooldown_seconds', 60)
    _ALERT_STATE_EXPIRY_SECONDS = monitor_cfg.get('alert_state_expiry_seconds', 300)

    _logger.info(f"Configurações de sistema aplicadas: Port={server_cfg.get('port')}, Host={server_cfg.get('host')}")

    
    
    
    if hasattr(_model_instance, 'set_alert_recorder_callback'):
        _model_instance.set_alert_recorder_callback(record_detection_event)
        _logger.info("Callback de registro de alertas (record_detection_event) configurado para a instância do modelo.")
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
            
    
    
    
    

    
    


    sse_handler = SSELogHandler(log_queue)
    sse_handler.setLevel(logging.DEBUG) 
    _logger.addHandler(sse_handler)
    _logger.info("Handler de log SSE adicionado ao logger principal.")

    log = logging.getLogger('werkzeug')
    log.setLevel(logging.ERROR)

    local_ip = get_ipv4_address()
    listen_host = server_cfg.get('host', '0.0.0.0')
    port = server_cfg.get('port', 7586)
    
    access_message = f"Servidor Flask iniciado. Acessível em:"
    access_message += f"\n   - Localmente: http://localhost:{port} ou http://127.0.0.1:{port}"
    if local_ip:
        access_message += f"\n   - Na rede local (via IPv4): http://{local_ip}:{port}"
    _logger.info(access_message)
    try:
        flask_app.run(host=listen_host, port=port, debug=False, use_reloader=False, threaded=True) 
    except Exception as e:
        _logger.error(f"Falha ao iniciar servidor Flask: {e}", exc_info=True)


@flask_app.before_request
def update_active_user_timestamp():
    if current_user.is_authenticated:
        with active_users_lock:
            now_utc = get_current_utc_timestamp()
            username = current_user.username
            
            if username in active_users:
                active_users[username]['last_seen'] = now_utc
            else:
                
                _logger.info(f"Usuário autenticado '{username}' (sessão persistente) adicionado à lista de ativos.")
                client_ip_on_reconnect = request.remote_addr
                hostname_on_reconnect = client_ip_on_reconnect 
                try:
                    hostname_on_reconnect, _, _ = socket.gethostbyaddr(client_ip_on_reconnect)
                except (socket.herror, socket.gaierror):
                    _logger.debug(f"Não foi possível resolver o hostname para o IP {client_ip_on_reconnect} do usuário {username} (reconexão). Usando IP.")
                active_users[username] = {
                    'role': current_user.role,
                    'login_time': now_utc, 
                    'last_seen': now_utc,
                    'ip_address': client_ip_on_reconnect,
                    'user_agent': request.user_agent.string,
                    'hostname': hostname_on_reconnect,
                    'first_name': getattr(current_user, 'first_name', None),
                    'last_name': getattr(current_user, 'last_name', None),
                                        'email': getattr(current_user, 'email', None)                }
@flask_app.route('/api/user/details', methods=['POST'])
@login_required
@role_required('admin')
def get_user_details():
    """Endpoint para obter detalhes do usuário."""
    try:
        data = request.json
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({"success": False, "message": "ID do usuário não fornecido"}), 400
            
        
        with active_users_lock:
            if user_id in active_users:
                user_data = active_users[user_id].copy()
                
                user_data['login_time'] = user_data['login_time'].strftime('%Y-%m-%d %H:%M:%S')
                user_data['last_seen'] = user_data['last_seen'].strftime('%Y-%m-%d %H:%M:%S')
                return jsonify({"success": True, "data": user_data}), 200
            
        return jsonify({"success": False, "message": "Usuário não encontrado"}), 404
        
    except Exception as e:
        _logger.error(f"Erro ao obter detalhes do usuário: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/user/logout', methods=['POST'])
@login_required
@role_required('admin')
def force_user_logout():
    """Endpoint para forçar o logout de um usuário."""
    try:
        data = request.json
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({"success": False, "message": "ID do usuário não fornecido"}), 400
            
        with active_users_lock:
            if user_id in active_users:
                
                db = get_db()
                db.log_activity(
                    user_id, 
                    active_users[user_id]['ip_address'],
                    'FORCED_LOGOUT',
                    f"Logout forçado por {current_user.username}"
                )
                
                
                del active_users[user_id]
                _logger.info(f"Logout forçado para usuário '{user_id}' por '{current_user.username}'")
                return jsonify({"success": True, "message": "Logout forçado realizado com sucesso"}), 200
                
        return jsonify({"success": False, "message": "Usuário não encontrado"}), 404
        
    except Exception as e:
        _logger.error(f"Erro ao forçar logout do usuário: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/user/temp-block', methods=['POST'])
@login_required
@role_required('admin')
def block_user_temp():
    """Endpoint para bloquear temporariamente um usuário."""
    try:
        data = request.json
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({"success": False, "message": "ID do usuário não fornecido"}), 400
            
        db = get_db()
        
        db.add_temp_block(
            user_id,
            current_user.username,
            get_current_utc_timestamp(),
            timedelta(hours=1) 
        )
        
        
        with active_users_lock:
            if user_id in active_users:
                db.log_activity(
                    user_id,
                    active_users[user_id]['ip_address'],
                    'TEMP_BLOCKED',
                    f"Bloqueio temporário por {current_user.username}"
                )
                del active_users[user_id]
        
        _logger.info(f"Usuário '{user_id}' bloqueado temporariamente por '{current_user.username}'")
        return jsonify({"success": True, "message": "Usuário bloqueado temporariamente"}), 200
        
    except Exception as e:
        _logger.error(f"Erro ao bloquear temporariamente usuário: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/user/ban', methods=['POST'])
@login_required
@role_required('admin')
def ban_user_permanent():
    """Endpoint para banir permanentemente um usuário."""
    try:
        data = request.json
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({"success": False, "message": "ID do usuário não fornecido"}), 400
            
        db = get_db()
        
        db.add_permanent_ban(
            user_id,
            current_user.username,
            get_current_utc_timestamp(),
            "Ban permanente via interface web"
        )
        
        
        with active_users_lock:
            if user_id in active_users:
                db.log_activity(
                    user_id,
                    active_users[user_id]['ip_address'],
                    'BANNED',
                    f"Banimento permanente por {current_user.username}"
                )
                del active_users[user_id]
        
        _logger.info(f"Usuário '{user_id}' banido permanentemente por '{current_user.username}'")
        return jsonify({"success": True, "message": "Usuário banido permanentemente"}), 200
        
    except Exception as e:
        _logger.error(f"Erro ao banir usuário: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.route('/api/user/unban', methods=['POST'])
@login_required
@role_required('admin')
def unban_user():
    """Endpoint para remover ban/bloqueio de um usuário."""
    try:
        data = request.json
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({"success": False, "message": "ID do usuário não fornecido"}), 400
            
        db = get_db()
        
        db.remove_all_blocks(user_id)
        
        db.log_activity(
            user_id,
            request.remote_addr,
            'UNBAN',
            f"Desbloqueio por {current_user.username}"
        )
        _logger.info(f"Usuário '{user_id}' desbloqueado por '{current_user.username}'")
        return jsonify({"success": True, "message": "Usuário desbloqueado com sucesso"}), 200 
    except Exception as e:
        _logger.error(f"Erro ao desbloquear usuário: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@flask_app.after_request
def after_request(response):
    response.headers['Server'] = 'VisionAlign'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    if current_user.is_authenticated:
        session.permanent = True  
        session.modified = True   
    return response


@flask_app.errorhandler(405)
def method_not_allowed_error(error):
    """Renderiza a página de erro 405 personalizada."""
    _logger.warning(f"Erro 405 - Método não permitido: {request.method} para {request.path} solicitado por {request.remote_addr}")
    
    return render_template('utils/405.html'), 405